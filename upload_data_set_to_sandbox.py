from openpyxl import load_workbook
from os import getcwd
from obp_python.config import logger
from obp_python.createAuthorityDataRequestInstanceAtBank import create_authority_data_request_instance_at_bank
from  obp_model.constants import BANK_TYPE, BANK_CODE
from sb.bank import create_bank_from_row, create_bank_attribute_from_row
from sb.customer import create_customer_from_row
from sb.account import create_account_from_row
from sb.product import create_product_from_row
from sb.branch import create_branch_from_row
from sb.authority_data_request import create_authority_data_request_from_row
from createAllBankEntitlementsForCurrentUser import create_bank_entitlements_for_user

date_format = '%Y-%m-%d'
data_source = f'{getcwd()}/resources/sb_import.xlsx'

wb = load_workbook(data_source)
banks = wb["Bancos"]
customers = wb['Cliente']
accounts = wb['Cuenta']
products = wb['Productos']
branches = wb['Sucursal-Location']
authority_requests = wb['Authority Data Request']


for row in banks.iter_rows(min_row=2, max_col=banks.max_column, max_row=banks.max_row):
	if row[0].value is None:
		continue
	logger.debug(f"start banks: {row[0].row}")
	bank_id = row[BANK_CODE].value.strip()
	logger.debug(f'This is the bank_id: {bank_id}')
	create_bank_from_row(row)
	create_bank_entitlements_for_user(bank_id)
	create_bank_attribute_from_row(row)
	create_authority_data_request_instance_at_bank(bank_id)
	# again for the dynamic entity
	create_bank_entitlements_for_user(bank_id)


for row in customers.iter_rows(min_row=2, max_col=customers.max_column, max_row=customers.max_row):
	if row[0].value is None:
		continue
	logger.debug(f"start customer: {row[0].row}")
	create_customer_from_row(row)

for row in accounts.iter_rows(min_row=2, max_col=accounts.max_column, max_row=accounts.max_row):
	if row[0].value is None:
		continue
	logger.debug(f"start account: {row[0].row}")
	create_account_from_row(row, date_format)

for row in products.iter_rows(min_row=2, max_col=products.max_column, max_row=products.max_row):
	if row[0].value is None:
		continue
	logger.debug(f"start product: {row[0].row}")
	create_product_from_row(row)

for row in branches.iter_rows(min_row=2, max_col=branches.max_column, max_row=branches.max_row):
	if row[0].value is None:
		continue
	logger.debug(f"start branch: {row[0].row}")
	create_branch_from_row(row)

for row in authority_requests.iter_rows(min_row=2, max_col=authority_requests.max_column, max_row=authority_requests.max_row):
	if row[0].value is None:
		logger.debug(f'skipping branch line {row[0].row}')
		continue
	logger.debug(f"start authority data request: {row[0].row}")
	create_authority_data_request_from_row(row)