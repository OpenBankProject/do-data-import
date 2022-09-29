from openpyxl import load_workbook
from os import getcwd



#data_source = f'{getcwd()}/resources/sb_data_import_template_ADOPEM.xlsx'
#data_source = f'{getcwd()}/resources/DataPruebaBancoPopular_revision.xlsx'
#data_source = f'{getcwd()}/resources/APAP - v1.0.0 20220923 - sb_data_import_template.xlsx'
#data_source = f'{getcwd()}/resources/APAP - v1.0.0 20220923 - sb_data_import_template_revised.xlsx'
#data_source = f'{getcwd()}/resources/sb_data_import_template_CONFISA.xlsx'
#data_source = f'{getcwd()}/resources/sb_data_import_template_Scotiabank.xlsx'
data_sources = [
	f'{getcwd()}/resources/sb_data_import_template_ADOPEM.xlsx',
	f'{getcwd()}/resources/DataPruebaBancoPopular_revision.xlsx',
	f'{getcwd()}/resources/APAP - v1.0.0 20220923 - sb_data_import_template.xlsx',
	# f'{getcwd()}/resources/APAP - v1.0.0 20220923 - sb_data_import_template_revised.xlsx',
	f'{getcwd()}/resources/sb_data_import_template_CONFISA.xlsx',
	f'{getcwd()}/resources/sb_data_import_template_Scotiabank.xlsx'
]
for data_source in data_sources:
	wb = load_workbook(data_source)
	banks = wb["Bancos"]
	customers = wb['Cliente']
	accounts = wb['Cuenta']
	products = wb['Productos']


	all_accounts = 0
	single_accounts = []
	duplicated_accounts = []
	all_products = 0
	product_only_accounts = []
	for row in accounts.iter_rows(min_row=2, max_col=accounts.max_column, max_row=accounts.max_row):
		if row[0].value is None:
			continue
		all_accounts += 1
		account_nr = row[2].value
		if account_nr in single_accounts:
			duplicated_accounts.append(account_nr)
		if account_nr not in single_accounts:
			single_accounts.append(account_nr)


	for row in products.iter_rows(min_row=2, max_col=products.max_column, max_row=products.max_row):
		if row[0].value is None:
			continue
		all_products += 1
		product_account_nr = row[5].value
		if product_account_nr not in single_accounts:
			product_only_accounts.append(product_account_nr)

	print(f'datasource: {data_source.rsplit("/")[-1]}')
	print(f"{len(duplicated_accounts)} duplicated accounts in cuenta out of {all_accounts}")
	#print(duplicated_accounts)

	print(f"{len(product_only_accounts)} products (productos) that have no account in  accounts (cuenta) out of {all_products} ")
	#print(product_only_accounts)